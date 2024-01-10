import errno
from distutils.dir_util import copy_tree
import os
from pathlib import Path
import shutil
import sqlite3

import openpyxl.workbook.protection
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection

import License_db as db
import login
from variables import path_db, path_server_backup, path_start_backup_txt_file, path_start_backup_txt_folder, path_xlsx_out
import get_messages as mes
from check_funcs import empty_or_not, check_path


from_directory = rf'{path_db}'
to_srv_directory = rf'{path_server_backup}'


def get_data_from_db():
    if db.check_db():
        data = ()
        with sqlite3.connect(path_db + '/Licenses.sqlite') as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM licenses ORDER BY name ASC")
            data = (row for row in cursor.fetchall())
            cursor.close()
            return data
    else:
        return 'Таблица не существует!'


def export_data_to_xlsx():
    data = list(get_data_from_db())
    if data == 'Таблица не существует!':
        return 'Таблица не существует!'
    else:
        # Создаем новую книгу
        try:
            wb = Workbook()
        except Exception as e:
            return f'Ошибка создания книги:\n{e}'

        try:
            # Создаем лист для данных
            wb.create_sheet('Лицензии', 0)
            ws = wb['Лицензии']

            # Добавляем заголовок таблицы в список
            header = ('№ п/п', 'ID', 'Название', 'Серийный номер', ' Шт.', 'Дата получения', 'Место установки', 'Дата установки')
            data.insert(0, header)

            # Заполняем данные
            counter = 0
            for row in data:
                temp = list(row)
                if counter != 0:
                    temp.insert(0, counter)
                ws.append(temp)
                counter += 1

            # Границы для ячеек и выравнивание текста
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Обрабатываем каждую заполненную ячейку
            for row in range(ws.min_row, ws.max_row + 1):
                for col in range(ws.min_column, ws.max_column + 1):
                    if row == 1:
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center",
                                                                           wrap_text=True)
                        ws.cell(row=row, column=col).border = thin_border
                    else:
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center",
                                                                           wrap_text=True)
                        ws.cell(row=row, column=col).border = thin_border

            # Выставляем для колонок ширину
            ws.column_dimensions[get_column_letter(1)].width = 5
            ws.column_dimensions[get_column_letter(2)].width = 5
            ws.column_dimensions[get_column_letter(3)].width = 42
            ws.column_dimensions[get_column_letter(4)].width = 54
            ws.column_dimensions[get_column_letter(5)].width = 6
            ws.column_dimensions[get_column_letter(6)].width = 12
            ws.column_dimensions[get_column_letter(7)].width = 49
            ws.column_dimensions[get_column_letter(8)].width = 15

            # Защищаем лист
            ws = wb['Лицензии']
            ws.protection.set_password(str(login.get_curent_pass()))

            # Скрываем лист
            ws.sheet_state = 'hidden'

            # Защищаем книгу
            wb.security = WorkbookProtection(workbookPassword=str(login.get_curent_pass()), lockWindows=True,
                                             lockStructure=True)

            ws = wb['Sheet']
            ws["A1"] = 'Чтобы увидеть список лицензий:'
            ws["A2"] = '        - разрешите редактирование'
            ws["A3"] = '        - снимите защиту с книги'
            ws["A4"] = '        (раздел "Рецензирование" с использованием текущего пароля в ПО "Лицензии")'
            ws["A5"] = '        - ПКМ по названию листа "Sheet"'
            ws["A6"] = '        - показать - "Лицензии"'
            ws.column_dimensions[get_column_letter(1)].width = 100

            # Сохраняем документ
            wb.save(path_xlsx_out)

            return True
        except Exception as e:
            return f'Ошибка создания файла xlsx:\n{e}'



# Проверка строки на русские символы
def check_rus(row):
    clear_row = ''
    russian = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'
    flag_rus = False
    rus_short = 'аекнорстхьАЕВКНОРСТХЬ'
    replace_dic = {
        'е': 'e',
        'а': 'a',
        'к': 'k',
        'н': 'h',
        'о': 'o',
        'р': 'p',
        'с': 'c',
        'т': 't',
        'х': 'x',
        'ь': 'b',
        'А': 'A',
        'Е': 'E',
        'В': 'B',
        'К': 'K',
        'Н': 'H',
        'О': 'O',
        'Р': 'P',
        'С': 'C',
        'Т': 'T',
        'Х': 'X',
        'Ь': 'B',
    }
    flag_rus_short = False
    rus_items = []
    show_rus_items = ''
    id_counter = 0
    for item in row:
        if item in russian:
            flag_rus = True
        if item in rus_short:
            rus_items.append((item, id_counter))
            flag_rus_short = True
        id_counter += 1
    if len(rus_items) > 0:
        temp_list = []
        for item in rus_items:
            temp_list.append(item[0])
        show_rus_items = ','.join(temp_list)
    if flag_rus:
        mes.warning('Русские символы', 'В строке "Серийный номер" обнаружены русские символы!')
        if flag_rus_short:
            result = mes.ask('Русские символы', 'Попытаться заменить русские символы из перечня "АЕВКНОРСТХЬ"?')
            if result == True:
                char_list = list(row)
                for item in rus_items:
                    new_ch = replace_dic[item[0]]
                    char_list[item[1]] = new_ch
                clear_row = ''.join(char_list)
                mes.info('Результат обработки строки', f'Обнаружено символов для замены: {len(rus_items)} - {show_rus_items}.\n'
                                                       f'Строка до - "{row}".\n'
                                                       f'Строка после - "{clear_row}".')
                return clear_row
            else:
                clear_row = row
                return clear_row
        else:
            clear_row = row
            return clear_row
    else:
        clear_row = row
        return clear_row


def clear_folder(path):
    import os, shutil
    folder = path
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
                print(f'Удалил {file_path}')
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
                print(f'Удалил {file_path}')
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))


def clear_old_backups(root, pass_path):
    paths = sorted(Path(root).iterdir(), key=os.path.getmtime)
    print(paths)
    print(pass_path)
    if len(paths) >= 2:
        for path in paths[:len(paths) - 2]:
            if str(path) == str(pass_path):
                print(f'Пропустили {path}')
                continue
            else:
                try:
                    print(f'Удаляем: {path}')
                    shutil.rmtree(path)
                except:
                    mes.warning('Удаление старой копии', f'При попытке удаления старой копии файлов'
                                                         f' произошла ошибка!\nУдаляли папку {path}')


# Резервное копирование по заданному пути
# на данном этапе изменение пути для РК не реализовано, но предусмотрено с необходимыми проверками
def backup_bd(extra_path):
    from datetime import datetime
    now_date = datetime.now().date().strftime("%d.%m.%Y")
    final_path = ''

    if extra_path == '':
        final_path = os.path.join(to_srv_directory, now_date)
    else:
        final_path = os.path.join(extra_path, now_date)
    root_dir = final_path.replace(now_date, '')
    try:
        if check_path(from_directory):
            if check_path(final_path):
                if empty_or_not(root_dir) is not None:
                    clear_old_backups(root_dir, final_path)
                if empty_or_not(final_path) is not None:
                    if mes.ask('Проверка пути для копирования',
                               f'Внимание! Конечная папка содержит файлы. Очистить её и продолжить копирование?\n\n{final_path}'):
                        clear_folder(final_path)
                        result = copy_tree(from_directory, final_path)
                        mes.info('Резервное копирование файлов',
                                 f'Успешно скопировано файлов: {len(result)}.\n\nСкопированы в: "{final_path}".')
                    else:
                        mes.error('Резервное копирование файлов', f'Отмена операции пользователем!')
                else:
                    result = copy_tree(from_directory, final_path)
                    mes.info('Резервное копирование файлов',
                             f'Успешно скопировано файлов: {len(result)}.\n\nСкопированы в: "{final_path}".')
            else:
                os.makedirs(final_path, exist_ok=True)
                result = copy_tree(from_directory, final_path)
                mes.info('Резервное копирование файлов',
                         f'Успешно скопировано файлов: {len(result)}.\n\nСкопированы в: "{final_path}".')
        else:
            mes.error('Резервное копирование файлов', f'Ошибка: путь с шаблонами не существует!\n\n{from_directory}')
    except OSError as exc:
        print(exc, exc.errno)
        # File already exist
        if exc.errno == errno.EEXIST:
            shutil.copy(from_directory, final_path)
        # The dirtory does not exist
        if exc.errno == errno.ENOENT:
            shutil.copy(from_directory, final_path)
        else:
            raise
    except Exception as e:
        mes.error('Резервное копирование БД', f'Ошибка: повторное копирование вызывает ошибку пути! Перезапустите программу и повторите попытку.\n\n----------\nОшибка:\n{e}')


# Восстановление файлов из последней резервной копии с сервера
def get_backup(root_func, status):
    # Куда копируем
    backup_to_path = from_directory
    # Папка с резервными копиями
    backup_root_dir = path_server_backup
    # Папка последней резервной копии
    backup_from_path = ''

    if check_path(backup_root_dir):
        if empty_or_not(backup_root_dir) is not None:
            paths = sorted(Path(backup_root_dir).iterdir(), key=os.path.getmtime)
            backup_from_path = str(paths[len(paths) - 1])
            print(backup_from_path)
            if not check_path(backup_to_path):
                os.makedirs(backup_to_path, exist_ok=True)
            if empty_or_not(backup_to_path) is not None:
                print(os.listdir(backup_to_path))
                if mes.ask('Проверка пути для копирования',
                           f'Внимание! Конечная папка содержит файлы. Очистить её и продолжить копирование?\n\n{backup_to_path}'):
                    print(f'Закрываем соединение!')
                    db = sqlite3.connect(path_db + '/Licenses.sqlite')
                    db.close()
                    with open(path_db + '/Licenses.sqlite', 'w+') as file:
                        file.close()
                    print(f'Закрыли соединение!')
                    print(f'Очищаем {backup_to_path}')
                    clear_folder(backup_to_path)
                    print(f'Копируем из {backup_from_path} в {backup_to_path}')
                    try:
                        result = copy_tree(backup_from_path, backup_to_path)
                        mes.info('Резервное копирование файлов',
                                 f'Успешно скопировано файлов: {len(result)}.\n\nСкопированы в: "{backup_to_path}".')
                    except Exception as e:
                        if str(e.__class__) == "<class 'distutils.errors.DistutilsFileError'>":
                            mes.error('Ошибка резервного копирования', 'Внимание!\n\nПри копировании не удалось получить'
                                                                       ' доступ к файлу БД для его замены.'
                                                                       '\n\nПерезапустите программу и сразу же '
                                                                       'попробуйте повторить копирование!')
                            if status == 1:
                                write_backup_file(1)
                            if root_func != '':
                                root_func()
                            else:
                                exit()
                        else:
                            mes.error('Ошибка резервного копирования',
                                      f'Внимание!\n\nПри копировании произошла непредвиденная ошибка!\n\nОшибка:\n{e}')
                else:
                    mes.error('Резервное копирование файлов', f'Отмена операции пользователем!')
            else:
                print(f'Копируем из {backup_from_path} в {backup_to_path}')
                try:
                    result = copy_tree(backup_from_path, backup_to_path)
                    mes.info('Резервное копирование файлов',
                             f'Успешно скопировано файлов: {len(result)}.\n\nСкопированы в: "{backup_to_path}".')
                except Exception as e:
                    if str(e.__class__) == "<class 'distutils.errors.DistutilsFileError'>":
                        mes.error('Ошибка резервного копирования', 'Внимание!\n\nПри копировании не удалось получить'
                                                                   ' доступ к файлу БД для его замены.'
                                                                   '\n\nПерезапустите программу и сразу же '
                                                                   'попробуйте повторить копирование!')
                        if status == 1:
                            write_backup_file(1)
                        if root_func != '':
                            root_func()
                        else:
                            exit()
                    else:
                        mes.error('Ошибка резервного копирования',
                                  f'Внимание!\n\nПри копировании произошла непредвиденная ошибка!\n\nОшибка:\n{e}')


# Создание пути
def create_path(new_path):
    try:
        os.makedirs(new_path, exist_ok=True)
    except Exception as e:
        mes.error('Создание пути', f'Ошибка создания пути!\n\nОшибка: [{e}]')


# Создать файл для начала резервного копирования
def write_backup_file(status):
    print(status)
    if check_path(path_start_backup_txt_file):
        os.remove(path_start_backup_txt_file)
        with open(path_start_backup_txt_file, 'w') as file:
            if status == 1:
                print('Записал True')
                file.write('True')
                file.close()
            else:
                print('Записал False')
                file.write('False')
                file.close()
    else:
        create_path(path_start_backup_txt_folder)
        with open(path_start_backup_txt_file, 'w') as file:
            if status == 1:
                print('Записал True')
                file.write('True')
                file.close()
            else:
                print('Записал False')
                file.write('False')
                file.close()


# Проверка необходимости резервного восстановления при запуске
def check_backup_status():
    print(f'Проверяю условие check_path(path_start_backup_txt_file) {check_path(path_start_backup_txt_file)}')
    if check_path(path_start_backup_txt_file):
        with open(path_start_backup_txt_file, 'r') as file:
            status = file.read()
            print(f'Прочитал статус {status}')
            file.close()
            print(f'Вернул статус {status}')
        os.remove(path_start_backup_txt_file)
        return status
    else:
        write_backup_file(0)
        with open(path_start_backup_txt_file, 'r') as file:
            status = file.read()
            print(f'Прочитал статус {status}')
            file.close()
            print(f'Вернул статус {status}')
        os.remove(path_start_backup_txt_file)
        return status
